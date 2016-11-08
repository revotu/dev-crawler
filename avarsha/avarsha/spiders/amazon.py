# -*- coding: utf-8 -*-
# @author: wanghaiyi

import urllib2
import re
from openpyxl import load_workbook

import scrapy.cmdline
from scrapy.selector import Selector

from avarsha_spider import AvarshaSpider


_spider_name = 'amazon'

class AmazonSpider(AvarshaSpider):
    name = _spider_name
    allowed_domains = ["amazon.com"]
    row_name = 1
    row_title = 1
    row_review = 1

    def __init__(self, *args, **kwargs):
        super(AmazonSpider, self).__init__(*args, **kwargs)

    def find_items_from_list_page(self, sel, item_urls):
        """parse items in category page"""

        base_url = 'http://www.amazon.com'
        items_xpath = ('//h3[@class="newaps"]/a/@href | //div'
            '/a[@class="a-link-normal a-text-normal"]/@href')

        # don't need to change this line
        return self._find_items_from_list_page(
            sel, base_url, items_xpath, item_urls)

    def find_nexts_from_list_page(self, sel, list_urls):
        """find next pages in category url"""

        base_url = 'http://www.amazon.com'
        nexts_xpath = ('//span[@class="pagnLink"]/a/@href | //a'
            '[@id="pagnNextLink"]/href')
        nexts = sel.xpath(nexts_xpath).extract()
        requests = []
        for path in nexts:
            list_url = path
            if path.find(base_url) == -1:
                list_url = base_url + path
            list_urls.append(list_url)
            request = scrapy.Request(list_url, callback=self.parse)
            requests.append(request)
        return requests

    def __remove_escape(self, content):
        content = content.replace('\\\"' , '"')
        content = content.replace('\\n' , '')
        content = content.replace('\\/' , '/')
        return content

    def _extract_url(self, sel, item):
        item['url'] = sel.response.url

    def _extract_title(self, sel, item):
        title_xpath = '//h1[@id="title"]/span/text()'
        data = sel.xpath(title_xpath).extract()
        if len(data) != 0:
            item['title'] = data[0].strip()

    def _extract_store_name(self, sel, item):
        item['store_name'] = 'Amazon'

    def _extract_brand_name(self, sel, item):
        brand_name_xpath = '//a[@id="brand"]/text()'
        data = sel.xpath(brand_name_xpath).extract()
        if len(data) != 0:
            item['brand_name'] = data[0].strip()
        else:
            brand_name_xpath = ('//div[@id="brand-teaser-story"]/'
                'span[@class="a-text-bold"]/text()')
            data = sel.xpath(brand_name_xpath).extract()
            if len(data) != 0:
                for i in range(len(data)):
                    if data[i].find('About '):
                        item['brand_name'] = data[i].replace('About ', '')


    def _extract_sku(self, sel, item):
        item['sku'] = sel.response.url[sel.response.url.find('/dp/') + len('/dp/'):sel.response.url.rfind('/')]

    def _extract_features(self, sel, item):
        features_xpath = '//div[@id="feature-bullets"]/ul/li/span/text()'
        data = sel.xpath(features_xpath).extract()
        if len(data) > 0:
            item['features'] = '<br>'.join(data)

    def _extract_description(self, sel, item):
        return
        description_tmp = ''
        description_xpath = '//div[@id="productDescription"]/p'
        description_detail_xpath = '//ul[@class="a-vertical a-spacing-none"]'
        data = sel.xpath(description_xpath).extract()
        if len(data) != 0:
            description_tmp = data[0]
        data = sel.xpath(description_detail_xpath).extract()
        if len(data) != 0:
            description_tmp += data[0]
        item['description'] = description_tmp

    def _extract_size_chart(self, sel, item):
        pass

    def _extract_color_chart(self, sel, item):
        pass

    def _extract_image_urls(self, sel, item):
        dir = 'amazon'
        img_reg = re.compile(r',"hiRes":"(.+?)","thumb"')
        data = img_reg.findall(sel.response.body)
        if len(data) > 0:
            item['image_urls'] = [ img + '?index=' + str(index + 1) + '&sku=' + item['sku'] + '&dir=' + dir for index,img in enumerate(data)]

    def _extract_colors(self, sel, item):
        return
        color_list = []
        for line in sel.response.body.split('\n'):
            idx1 = line.find('\"variationValues\" : ')
            if line.find('color_name') != -1:
                if idx1 != -1:
                    idx_num = line.find('[\"' , idx1 + len('\"variationValues\" : '))
                    idx2 = '  ,' + line[idx_num + 1:]
                    idx_num2 = idx2.find('],')
                    idx3 = idx2[:idx_num2 - 1]
                    while idx3.rfind(',') != -1:
                        color_tmp = idx3[ idx3.rfind(',') + 2:].strip()
                        color_list.append(unicode(color_tmp))
                        idx3 = idx3[:idx3.rfind(',') - 1]
        if len(color_list) != 0:
            item['colors'] = color_list

    def _extract_sizes(self, sel, item):
        return
        size_xpath = '//option[@data-a-css-class="dropdownAvailable"]/@data-a-html-content'
        data = sel.xpath(size_xpath).extract()
        if len(data) != 0:
            item['sizes'] = data

    def _extract_stocks(self, sel, item):
        pass

    def _extract_price(self, sel, item):
        price_xpath = ('//span[@id="priceblock_ourprice"]/text() | '
            '//span[@id="priceblock_saleprice"]/text()')
        data = sel.xpath(price_xpath).extract()
        if len(data) != 0:
            if data[0].rfind('-') != -1:
                data_tmp = data[0][data[0].rfind('$') + 1:]
                item['price'] = self._format_price('USD', data_tmp.replace('$', ''))
            else:
                item['price'] = self._format_price('USD', data[0].replace('$', ''))

    def _extract_list_price(self, sel, item):
        return
        list_price_xpath = ('//tr/td[@class="a-span12 a-color-secondary '
            'a-size-base a-text-strike"]/text()')
        data = sel.xpath(list_price_xpath).extract()
        if len(data) != 0:
            item['list_price'] = self._format_price('USD', data[0].replace('$', ''))

    def _extract_low_price(self, sel, item):
        return
        price_xpath = ('//span[@id="priceblock_ourprice"]/text() | '
            '//span[@id="priceblock_saleprice"]/text()')
        data = sel.xpath(price_xpath).extract()
        if len(data) != 0:
            if data[0].rfind('-') != -1:
                data_tmp = data[0][:data[0].rfind('-')].replace('$' , '')
                item['low_price'] = self._format_price('USD', data_tmp)

    def _extract_high_price(self, sel, item):
        return
        price_xpath = ('//span[@id="priceblock_ourprice"]/text() | '
            '//span[@id="priceblock_saleprice"]/text()')
        data = sel.xpath(price_xpath).extract()
        if len(data) != 0:
            if data[0].rfind('-') != -1:
                data_tmp = data[0][data[0].rfind('-') + 2:].replace('$' , '')
                item['high_price'] = self._format_price('USD', data_tmp)
    def _extract_is_free_shipping(self, sel, item):
        pass

    def _extract_review_count(self, sel, item):
        pass

    def _extract_review_rating(self, sel, item):
        pass

    def _extract_best_review_rating(self, sel, item):
        pass

    def _extract_review_list(self, sel, item):
        return
        sel = Selector(sel.response)
        review_count = 0
        review_rating = 0
        err = 0
        review_list = []
        first_review_xpath = '//div/a[@class="a-link-emphasis a-nowrap"]/@href'
        first_review_url = sel.xpath(first_review_xpath).extract()
        if len(first_review_url) != 0:
            for err in range(3):
                try:
                    content = urllib2.urlopen(first_review_url[0]).read()
                    content = self.__remove_escape(content)
                    err = 3
                except:
                    continue
            sel = Selector(text=content)
            review_count_xpath = '//span[@class="a-size-medium a-text-beside-button totalReviewCount"]/text()'
            data = sel.xpath(review_count_xpath).extract()
            if len(data) != 0:
                review_list = []
                review_count = int(data[0].strip().replace(',', ''))
                review_rating_xpath = ('//span[@class='
                    '"arp-rating-out-of-text"]/text()')
                data = sel.xpath(review_rating_xpath).extract()
                if len(data) != 0:
                    if data[0].find('out of') != -1:
                        review_rating = float(data[0][:data[0].find('out') - 1].strip())
                    else:
                        review_rating_xpath = ('//div[@class='
                            '"a-row averageStarRatingNumerical"]/span/text()')
                        data = sel.xpath(review_rating_xpath).extract()
                        if len(data) != 0:
                            review_rating = float(data[0][:data[0].find('out') - 1].strip())
                for i in range(0, (review_count / 10) + 1):
                    if i != 0:
                        list_review_url = []
                        next_url_xpath = '//li[@class="a-last"]/a/@href'
                        review_url = sel.xpath(next_url_xpath).extract()
                        if len(review_url) != 0:
                            try:
                                list_review_url.append('http://www.amazon.com' + review_url[0])
                                content = urllib2.urlopen(list_review_url[0]).read()
                                content = self.__remove_escape(content)
                                sel = Selector(text=content)
                            except:
                                continue
                    review_name_xpath = '//span/a[@class="a-size-base a-link-normal author"]/text()'
                    names = sel.xpath(review_name_xpath).extract()
                    review_date_xpath = ('//div/span[@class='
                        '"a-size-base a-color-secondary review-date"]/text()')
                    dates = sel.xpath(review_date_xpath).extract()
                    rating_xpath = ('//div[@id="cm_cr-review_list"]//a[@class="a-link-normal"]'
                        '/i/span[@class="a-icon-alt"]/text()')
                    ratings_tmp = sel.xpath(rating_xpath).extract()
                    for k in range(len(ratings_tmp)):
                        ratings_tmp[k] = float(ratings_tmp[k][:ratings_tmp[k].find('out') - 1].strip())
                    ratings = ratings_tmp
                    titles_xpath = ('//a[@class = "a-size-base a-link-normal '
                        'review-title a-color-base a-text-normal a-text-bold"]/'
                        'text() | //a[@class="a-size-base a-link-normal review-'
                        'title a-color-base a-text-bold"]/text()')
                    titles = sel.xpath(titles_xpath).extract()
                    review_content_xpath = ('//div[@class="a-row review-data"]/'
                        'span/text()')
                    review_contents = sel.xpath(review_content_xpath).extract()
                    for j in range(len(names)):
                        while (len(review_contents) < len(names)):
                            review_contents.append("")
                        review_list.append({'rating':ratings[j],
                            'date':dates[j], 'name':names[j],
                            'title':titles[j], 'content':review_contents[j]})

                    xls_name = load_workbook('names.xlsx')
                    xls_title = load_workbook('titles.xlsx')
                    xls_review = load_workbook('reviews.xlsx')
                    table_name = xls_name.active
                    table_title = xls_title.active
                    table_review = xls_review.active
                    for index in range(len(names)):
                        if ratings[index] >= 4:
                            table_name.cell(row=self.row_name, column=1).value = names[index]
                            self.row_name += 1
                            table_title.cell(row=self.row_title, column=1).value = titles[index]
                            self.row_title += 1
                            table_review.cell(row=self.row_review, column=1).value = review_contents[index]
                            self.row_review += 1
                    xls_name.save('names.xlsx')
                    xls_title.save('titles.xlsx')
                    xls_review.save('reviews.xlsx')

                    print 'OK'
                    print '================================================='
                    print self.row_name
                    print '================================================='
                    print 'OK'

        if review_rating != 0:
            item['max_review_rating'] = 5
            item['review_rating'] = review_rating
        if review_count != 0:
            item['review_count'] = review_count
        if len(review_list) != 0:
            item['review_list'] = review_list

def main():
    scrapy.cmdline.execute(argv=['scrapy', 'crawl', _spider_name])

if __name__ == '__main__':
    main()
