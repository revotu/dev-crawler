# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html


import os.path
import rfc822
import time
import json
from cStringIO import StringIO
from PIL import Image
from sets import Set
from twisted.internet import defer

from scrapy.http import Request
from scrapy import log
from scrapy.contrib.pipeline.files import S3FilesStore
from scrapy.contrib.pipeline.images import ImagesPipeline
from scrapy.exceptions import DropItem
from boto.dynamodb2.exceptions import ItemNotFound

from avarsha.dynamo_db import DynamoDB
from openpyxl import load_workbook


class AvarshaPipeline(object):
    def __init__(self):
        self.dynamo_db = DynamoDB()

    def __build_product_availabity_item(self, product_item):
        if (product_item.get('colors') is None
            and product_item.get('sizes') is None
            and product_item.get('stocks') is None):
            return None

        pa_item = {}
        pa_item['product_id'] = product_item['product_id']
        if product_item.get('colors') != None:
            pa_item['colors'] = product_item['colors']
        if product_item.get('sizes') != None:
            pa_item['sizes'] = product_item['sizes']
        if product_item.get('stocks') != None:
            pa_item['stocks'] = product_item['stocks']
        pa_item['crawl_datetime'] = product_item['crawl_datetime']
        return pa_item

    def __build_product_sales_item(self, product_item):
        if (product_item.get('list_price') is None
            and product_item.get('price') is None
            and product_item.get('low_price') is None
            and product_item.get('high_price') is None
            and product_item.get('free_shipping') is None):
            return None

        ps_item = {}
        ps_item['product_id'] = product_item['product_id']
        if product_item.get('list_price') != None:
            ps_item['list_price'] = product_item['list_price']
        if product_item.get('price') != None:
            ps_item['price'] = product_item['price']
        if product_item.get('low_price') != None:
            ps_item['low_price'] = product_item['low_price']
        if product_item.get('high_price') != None:
            ps_item['high_price'] = product_item['high_price']
        if product_item.get('free_shipping') != None:
            ps_item['free_shipping'] = product_item['free_shipping']
        ps_item['crawl_datetime'] = product_item['crawl_datetime']
        return ps_item

    def __build_product_reviews_item(self, product_item):
        if (product_item.get('review_count') is None
            and product_item.get('max_review_rating') is None
            and product_item.get('review_rating') is None
            and product_item.get('review_list') is None):
            return None

        pr_item = {}
        pr_item['product_id'] = product_item['product_id']
        if product_item.get('review_count') != None:
            pr_item['review_count'] = product_item['review_count']
        if product_item.get('max_review_rating') != None:
            pr_item['max_review_rating'] = product_item['max_review_rating']
        if product_item.get('review_rating') != None:
            pr_item['review_rating'] = product_item['review_rating']
        if product_item.get('review_list') != None:
            pr_item['review_list'] = product_item['review_list']
        pr_item['crawl_datetime'] = product_item['crawl_datetime']
        return pr_item

    def process_item(self, item, spider):
#         if item.normalize_attributes() is False:
#             raise DropItem("Attributes format error in %s" % item)
#         self.__assert_necessary_attributes(item)

        if spider.settings['VERSION'] == 'DEV':
            #self.store(item)
            return item

        if spider.settings['CHROME_ENABLED'] is True:
            try:
                chrome_item = self.dynamo_db.table('chrome').get_item(url=item['url'])
                item['chrome_clicks'] = chrome_item['clicks']
            except ItemNotFound:
                log.msg('URL not in chrome log: %s' % item['url'] , log.DEBUG)

        product_id = item.get('product_id')
        old_item = None
        try:
            old_item = self.dynamo_db.table('products').get_item(product_id=product_id)
            item['collections'] = list(item['collections'])
            if old_item != item:
                item['updated_datetime'] = item['crawl_datetime']
                log.msg('Item updated since last crawl: %s' % product_id, log.DEBUG)
            else:
                # detect if it is crawled recently
                old_item['crawl_datetime'] = item['crawl_datetime']
                old_item.partial_save()
                log.msg('Item not changed: %s' % product_id, log.DEBUG)
                return item
        except ItemNotFound:
            log.msg('Find a new item: %s' % product_id, log.DEBUG)
            item['updated_datetime'] = item['crawl_datetime']

        # merge current collections to previous collections
        if old_item != None:
            new_collections = Set(old_item.get('collections')) \
                | Set(item.get('collections'))
        else:
            new_collections = Set(item.get('collections'))
        if len(new_collections) != 0:
            item['collections'] = list(new_collections)

        log.msg('Save item info to Dynamo DB: %s' % product_id, log.DEBUG)

        self.dynamo_db.table('products').put_item(data=item, overwrite=True)

        pa_item = self.__build_product_availabity_item(item)
        if pa_item != None:
            self.dynamo_db.table('product_availabity')\
                .put_item(data=pa_item, overwrite=True)

        ps_item = self.__build_product_sales_item(item)
        if ps_item != None:
            self.dynamo_db.table('product_sales')\
                .put_item(data=ps_item, overwrite=True)

        pr_item = self.__build_product_reviews_item(item)
        if pr_item != None:
            self.dynamo_db.table('product_reviews')\
                .put_item(data=pr_item, overwrite=True)

        return item

    def open_spider(self, spider):
        feeder = spider.feeder

        if spider.settings['VERSION'] == 'DEV':
            start_urls = ['https://www.amazon.com/Miusol-Womens-Vintage-Wedding-X-Large/dp/B01A6P0VLQ/ref=sr_1_14']
            
#             dir = os.path.dirname(os.path.realpath(__file__))
#             wb = load_workbook(os.path.join(dir,'..','..','1688-1014.xlsx'))
#             ws = wb.active
#             for i in range(1,277):
#                 start_urls.append(ws.cell(row = i,column = 1).value + '?row=' + str(i))
#             wb.save(os.path.join(dir,'..','..','1688-1014.xlsx'))
                
            feeder.init_test_feeds(start_urls)
        else:
            feeder.init_feeds(spider_name=spider.name,
                feed_type=spider.feed_type)

        # for collection and products relationship
        for feed in feeder.collection_feeds:
            collection = feed[0].strip()
            url = spider.convert_url(feed[2].strip())
            spider.start_urls.append(url)
            collection_set = feeder.collections(url)
            collection_set.add(collection)
            feeder.map_url_collections(url, collection_set)

    def close_spider(self, spider):
        return  # TODO: currently, we don't update crawl time
        if spider.settings['VERSION'] == 'RELEASE':
            feeder = spider.feeder
            feeder.update_next_crawl_datetime()

    def __assert_necessary_attributes(self, item):
        assert_fields = ('title', 'price', 'collections', 'image_urls', 'sku')
        for field in assert_fields:
            if item.get(field) is None:
                raise DropItem("Missing field [%s] in %s" % (field, item))
        if len(item.get('images')) == 0:
            raise DropItem("Download images error.")
    
    def init_to_excel(self , item):
        dir = os.path.dirname(os.path.realpath(__file__))
        wb = load_workbook(os.path.join(dir,'..','..','terms-products.xlsx'))
        ws = wb.active
        data = []
        data.append(item['referer'])
        data.append(item['url'])
        data.append(item['product_id'])
        ws.append(data)
        wb.save(os.path.join(dir,'terms-products.xlsx'))

    def store(self ,item):
        dir = os.path.dirname(os.path.realpath(__file__))
        wb = load_workbook(os.path.join(dir,'..','..','1688-1012.xlsx'))
        ws = wb.active
        ws.cell(row = int(item['sku']),column = 3).value = item['price']
        ws.cell(row = int(item['sku']),column = 4).value = item['sizes']
        ws.cell(row = int(item['sku']),column = 5).value = item['colors']
        print os.path.join(dir,'..','..','1688-1012.xlsx')
        wb.save(os.path.join(dir,'..','..','1688-1012.xlsx'))
        
    
    def store_to_excel(self , item):
        dir = os.path.dirname(os.path.realpath(__file__))
        wb = load_workbook(os.path.join(dir,'..','..','products.xlsx'))
        ws = wb.active
        data = []
        data.append(item['sku'])
        data.append(item['product_id'])
        data.append('dress')
        data.append(item['title'])
        data.append(item['description'])
        data.append(item['price'][len('USD '):])
        data.append(item['list_price'][len('USD '):])
        data.append('1.5')
        data.append('2')
        data.append('Blushing Pink%%Champagne%%Iovry%%White')
        data.append('')
        data.append('')
        data.append('2%%4%%6%%8%%10%%12%%14%%16%%16W%%18W%%20W%%22W%%24W%%26W')
        
        data.append(item['features']['SILHOUETTE'].replace(', ','%%') if 'SILHOUETTE' in item['features'] else '')
        data.append(item['features']['HEMLINE / TRAIN'].replace(', ','%%') if 'HEMLINE / TRAIN' in item['features'] else '')
        data.append(item['features']['NECKLINE'].replace(', ','%%') if 'NECKLINE' in item['features'] else '')
        data.append(item['features']['SLEEVE LENGTH'].replace(', ','%%') if 'SLEEVE LENGTH' in item['features'] else '')
        data.append(item['features']['WAIST'].replace(', ','%%') if 'WAIST' in item['features'] else '')
        data.append(item['features']['SLEEVE'].replace(', ','%%') if 'Sleeve' in item['features'] else '')
        data.append(item['features']['FABRIC'].replace(', ','%%') if 'FABRIC' in item['features'] else '')
        data.append(item['features']['EMBELLISHMENT'].replace(', ','%%') if 'EMBELLISHMENT' in item['features'] else '')
        data.append('Yes')
        data.append('Yes')
        data.append(item['features']['BONING'].replace(', ','%%') if 'Boning' in item['features'] else '')
        data.append('')
        data.append(item['features']['BODY SHAPE'].replace(', ','%%') if 'BODY SHAPE' in item['features'] else '')
        data.append(item['features']['BELT FABRIC'].replace(', ','%%') if 'Belt Fabric' in item['features'] else '')
        data.append(item['features']['BACK STYLE'].replace(', ','%%') if 'Back Style' in item['features'] else '')
        data.append('')
        data.append(item['features']['TREND'].replace(', ','%%') if 'Trend' in item['features'] else '')
        data.append(item['features']['STYLE'].replace(', ','%%')  if 'STYLE' in item['features'] else '')
        data.append(item['features']['OCCASION'].replace(', ','%%') if 'Occasion' in item['features'] else '')
        data.append(item['features']['SEASON'].replace(', ','%%') if 'SEASON' in item['features'] else '')
        data.append(item['features']['WEDDING VENUES'].replace(', ','%%') if 'WEDDING VENUES' in item['features'] else '')
        data.append('5')
        data.append('789')
        
        ws.append(data)
        wb.save(os.path.join(dir,'products.xlsx'))
        
        if len(item['review_list']) != 0:
            wb = load_workbook(os.path.join(dir,'reviews.xlsx'))
            ws = wb.active
            for rev in item['review_list']:
                reviews = []
                reviews.append(item['sku'])
                reviews.append(rev['name'])
                reviews.append(rev['content'])
                
                ws.append(reviews)
                wb.save(os.path.join(dir,'reviews.xlsx'))

class AvarshaS3FilesStore(S3FilesStore):
    def __init__(self, *args, **kwargs):
        super(AvarshaS3FilesStore, self).__init__(*args, **kwargs)

    def stat_file(self, path, info):
        def _onsuccess(boto_key):
            checksum = boto_key.etag.strip('"')
            last_modified = boto_key.last_modified
            modified_tuple = rfc822.parsedate_tz(last_modified)
            modified_stamp = int(rfc822.mktime_tz(modified_tuple))
            return {'checksum': checksum,
                    'last_modified': modified_stamp,
                    'width': boto_key.metadata.width,
                    'height': boto_key.metadata.height}
        return self._get_boto_key(path).addCallback(_onsuccess)

class AvarshaImagePipeline(ImagesPipeline):
    def __init__(self, *args, **kwargs):
        self.STORE_SCHEMES['s3'] = AvarshaS3FilesStore
        super(AvarshaImagePipeline, self).__init__(*args, **kwargs)

    def media_to_download(self, request, info):
        def _onsuccess(result):
            if not result:
                return  # returning None force download

            last_modified = result.get('last_modified', None)
            if not last_modified:
                return  # returning None force download

            age_seconds = time.time() - last_modified
            age_days = age_seconds / 60 / 60 / 24
            if age_days > self.EXPIRES:
                return  # returning None force download

            referer = request.headers.get('Referer')
            log.msg(format='File (uptodate): Downloaded %(medianame)s from %(request)s referred in <%(referer)s>',
                    level=log.DEBUG, spider=info.spider,
                    medianame=self.MEDIA_NAME, request=request, referer=referer)
            self.inc_stats(info.spider, 'uptodate')

            checksum = result.get('checksum', None)
            width = result.get('width', None)
            height = result.get('height', None)
            return {'url': request.url, 'path': path, 'checksum': checksum, 'width': width, 'height': height}

        path = self.file_path(request, info=info)
        dfd = defer.maybeDeferred(self.store.stat_file, path, info)
        dfd.addCallbacks(_onsuccess, lambda _: None)
        dfd.addErrback(log.err, self.__class__.__name__ + '.store.stat_file')
        return dfd


    def media_downloaded(self, response, request, info):
        # extend width and height to item['images']
        default_meta = super(ImagesPipeline, self).\
            media_downloaded(response, request, info)
        (width, height) = self.get_image_sizes(response, request)
        return dict(default_meta.items()
            + {'width': width, 'height': height}.items())

    def get_image_sizes(self, response, request):
        orig_image = Image.open(StringIO(response.body))
        width, height = orig_image.size
        return (width, height)

    def file_path(self, request, response=None, info=None):
        ## start of deprecation warning block (can be removed in the future)
        def _warn():
            from scrapy.exceptions import ScrapyDeprecationWarning
            import warnings
            warnings.warn('ImagesPipeline.image_key(url) and file_key(url) methods are deprecated, '
                          'please use file_path(request, response=None, info=None) instead',
                          category=ScrapyDeprecationWarning, stacklevel=1)

        # check if called from image_key or file_key with url as first argument
        if not isinstance(request, Request):
            _warn()
            url = request
        else:
            url = request.url

        # detect if file_key() or image_key() methods have been overridden
        if not hasattr(self.file_key, '_base'):
            _warn()
            return self.file_key(url)
        elif not hasattr(self.image_key, '_base'):
            _warn()
            return self.image_key(url)
        ## end of deprecation warning block
        
        index = url[url.find('?index=') + len('?index='):url.find('&sku=')]
        sku = url[url.find('&sku=') + len('&sku='):url.find('&dir=')]
        dir = url[url.find('&dir=') + len('&dir='):]
        
        return '1688/%s/%s_(%s).jpg' % (dir,sku,index)